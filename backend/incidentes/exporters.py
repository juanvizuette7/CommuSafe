"""Exportadores de incidentes para el panel web."""

from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


COLUMNAS_INCIDENTES = [
    "Título",
    "Categoría",
    "Prioridad",
    "Estado",
    "Reportado por",
    "Fecha reporte",
    "Fecha cierre",
    "Atendido por",
    "Ubicación",
]


def _fecha_archivo():
    return timezone.localtime().strftime("%Y%m%d")


def _formatear_fecha(fecha):
    if not fecha:
        return ""
    return timezone.localtime(fecha).strftime("%d/%m/%Y %I:%M %p")


def _fila_incidente(incidente):
    return [
        incidente.titulo,
        incidente.get_categoria_display(),
        incidente.get_prioridad_display(),
        incidente.get_estado_display(),
        incidente.reportado_por.nombre_completo,
        _formatear_fecha(incidente.fecha_reporte),
        _formatear_fecha(incidente.fecha_cierre),
        incidente.atendido_por.nombre_completo if incidente.atendido_por else "",
        incidente.ubicacion_referencia or "",
    ]


def exportar_incidentes_excel(queryset):
    """Exporta el queryset filtrado de incidentes a un archivo XLSX."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Historial"

    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(color="FFFFFF", bold=True)
    worksheet.append(COLUMNAS_INCIDENTES)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for incidente in queryset:
        worksheet.append(_fila_incidente(incidente))

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, _ in enumerate(COLUMNAS_INCIDENTES, start=1):
        column_letter = get_column_letter(column_index)
        max_length = 14
        for cell in worksheet[column_letter]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value) + 2, 48))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.column_dimensions[column_letter].width = max_length

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename = f"commusafe_incidentes_{_fecha_archivo()}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _texto_filtros(filtros):
    etiquetas = {
        "categoria": "Categoría",
        "estado": "Estado",
        "prioridad": "Prioridad",
        "q": "Búsqueda",
    }
    activos = [
        f"{etiquetas.get(clave, clave)}: {valor}"
        for clave, valor in filtros.items()
        if str(valor or "").strip()
    ]
    return "Filtros aplicados: " + ", ".join(activos) if activos else "Filtros aplicados: ninguno"


def _footer(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    generado = timezone.localtime().strftime("%d/%m/%Y %I:%M %p")
    canvas.drawString(doc.leftMargin, 20, f"Generado el {generado}")
    canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, 20, f"Página {doc.page}")
    canvas.restoreState()


def exportar_incidentes_pdf(queryset, filtros):
    """Exporta el queryset filtrado de incidentes a un PDF tabular."""

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=24,
        leftMargin=24,
        topMargin=28,
        bottomMargin=34,
    )
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    normal_style = styles["BodyText"]
    normal_style.fontSize = 8
    normal_style.leading = 10

    elementos = [
        Paragraph("CommuSafe — Historial de incidentes", title_style),
        Paragraph(_texto_filtros(filtros), styles["Normal"]),
        Spacer(1, 12),
    ]

    data = [COLUMNAS_INCIDENTES]
    for incidente in queryset:
        data.append([Paragraph(str(valor), normal_style) for valor in _fila_incidente(incidente)])

    if len(data) == 1:
        data.append(["Sin resultados", "", "", "", "", "", "", "", ""])

    tabla = Table(
        data,
        repeatRows=1,
        colWidths=[120, 70, 58, 68, 105, 75, 75, 95, 110],
    )
    tabla.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A1A2E")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 7),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    elementos.append(tabla)

    doc.build(elementos, onFirstPage=_footer, onLaterPages=_footer)
    buffer.seek(0)

    filename = f"commusafe_incidentes_{_fecha_archivo()}.pdf"
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
